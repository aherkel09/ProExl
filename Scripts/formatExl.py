from openpyxl import Workbook, load_workbook
from datetime import datetime

class Pyxl():
    def __init__(self, template):
        self.today = datetime.now()
        self.date = self.today.strftime('%Y-%m-%d')
        self.template = template
        self.saveSheetAs = 'Templates/databaseRevisions' + str(self.date) + '.xlsx'
        self.row = 3
        #column indices and default data. '' entries will be overwritten by makeDuct.py
        self.activeColumns = {
            'Code':[1, ''], 'DivisionCode':[2, ''], 'SubDivisionCode':[3, ''],
            'ItemCode':[4, ''], 'DivisionDescription':[5, ''],
            'SubDivisionDescription':[6, ''], 'ItemDescription':[7, ''],
            'ItemUnit':[8, 'LF'], 'MaterialVendor':[14, 'RADCO'], 'Size':[19, 0],
            'MaterialCost':[20, 1.00], 'MaterialConversion':[21, 1.656], 'MaterialRounding':[22,'None'],
            'MaterialUnit':[23, 'LB'], 'MaterialWastepercentage':[24, 0], 'LaborCost':[27, 75],
            'LaborConversion':[28, 1.0], 'LaborRounding':[29, 'None'], 'LaborUnit':[30, 'HR'],
            'LaborWastepercentage':[31, 0], 'SubContractorConversion':[35, 1],
            'SubContractorRounding':[36, 'None'],'SubContractorWastepercentage':[38, 0],
            'EquipmentConversion':[42, 1], 'EquipmentRounding':[43, 'None'],
            'EquipmentWastepercentage':[45, 0], 'OtherConversion':[49, 1],
            'OtherRounding':[50, 'None'], 'OtherWastepercentage':[52, 0]
            }

    def trackRevisions(self):
        #tracks database revision number in text file.
        self.file = open('Templates/RevisionNum.txt', 'a')
        self.file.write('Database Revised ' + str(self.today) + '\n')
        self.file.close()
        
    def loadTemplateSpreadsheet(self, template):
        #loads ProEst Template excel sheet.
        self.templateWorkbook = load_workbook(template, data_only=True)
        self.sheetName = self.templateWorkbook.get_sheet_names()[0]
        self.templateWorksheet = self.templateWorkbook.get_sheet_by_name(self.sheetName)

    def createNewSpreadsheet(self):
        self.loadTemplateSpreadsheet(self.template)
        self.newWorkbook = Workbook()
        self.newWorksheet = self.newWorkbook.active
        self.newWorksheet.title = 'ProEst Template'
        self.getHeaders(self.templateWorksheet)
        
    def getHeaders(self, sheet):
        #gets headers from template worksheet
        self.headers = []
        for x in range(1, 100):
            self.headerCellValue = self.templateWorksheet.cell(row=1, column=x).value
            self.headers.append(self.headerCellValue)
        self.writeHeaders(self.headers)

    def writeHeaders(self, headers):
        #adds headers to new excel sheet.
        self.newWorksheet.append(headers)
        self.writeData()

    def writeData(self):
        #called from makeDuct.py. writes data from activeColumns to excel sheet.
        for col in self.activeColumns:
            self.listValues = self.activeColumns[col]
            self.columnIndex = self.listValues[0]
            self.data = self.listValues[1]
            self.newWorksheet.cell(row=self.row, column=self.columnIndex,
                                   value=self.data)

    def saveSpreadsheet(self):
        #saves new excel sheet.
        self.newWorkbook.save(self.saveSheetAs)
        self.trackRevisions()
